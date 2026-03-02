import secrets
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import List, Optional
from app.database import get_db
from app.models.project import Project
from app.models.user import User
from app.models.client import Client
from app.schemas.project import ProjectCreate, ProjectUpdate, ProjectOut
from app.services.auth import get_current_user

router = APIRouter(prefix="/api/projects", tags=["projects"])


STATUS_RU = {
    "PLANNING": "Планирование",
    "IN_PROGRESS": "В работе",
    "ON_HOLD": "На паузе",
    "COMPLETED": "Завершен",
    "CANCELLED": "Отменен",
}

PRIORITY_RU = {
    "LOW": "Низкий",
    "MEDIUM": "Средний",
    "HIGH": "Высокий",
    "URGENT": "Срочный",
}


@router.get("/", response_model=List[ProjectOut])
async def list_projects(db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    result = await db.execute(select(Project).order_by(Project.id.desc()))
    return [ProjectOut.model_validate(p) for p in result.scalars().all()]


@router.post("/", response_model=ProjectOut)
async def create_project(data: ProjectCreate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    project = Project(**data.model_dump(), owner_id=user.id)
    db.add(project)
    await db.flush()
    await db.refresh(project)
    return ProjectOut.model_validate(project)


@router.get("/export")
async def export_projects_excel(db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    """Export all projects to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    result = await db.execute(select(Project).order_by(Project.id.desc()))
    projects = result.scalars().all()

    # Get clients for name lookup
    clients_result = await db.execute(select(Client))
    clients = {c.id: c.company_name for c in clients_result.scalars().all()}

    # Get users for name lookup
    users_result = await db.execute(select(User))
    users = {u.id: u.name for u in users_result.scalars().all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Проекты"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["ID", "Название", "Описание", "Клиент", "Статус", "Приоритет",
               "Бюджет", "Потрачено", "Прогресс %", "Дата начала", "Дедлайн", "Ответственный"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data
    for row, p in enumerate(projects, 2):
        data = [
            p.id,
            p.name,
            p.description or "",
            clients.get(p.client_id, "—") if p.client_id else "—",
            STATUS_RU.get(p.status.value, p.status.value),
            PRIORITY_RU.get(p.priority.value, p.priority.value),
            f"{p.budget:,.0f} ₸",
            f"{p.spent:,.0f} ₸",
            f"{p.progress}%",
            p.start_date.strftime("%d.%m.%Y") if p.start_date else "",
            p.deadline.strftime("%d.%m.%Y") if p.deadline else "",
            users.get(p.owner_id, "—"),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Adjust column widths
    column_widths = [6, 30, 40, 25, 15, 12, 15, 15, 12, 12, 12, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=projects.xlsx"}
    )


@router.get("/shared/{share_token}", response_model=List[ProjectOut])
async def get_shared_projects(share_token: str, db: AsyncSession = Depends(get_db)):
    """Public access via share token - returns all projects for that token"""
    result = await db.execute(select(Project).where(Project.share_token == share_token))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Share link not found")
    # Return all projects (or you can return just the shared one)
    all_projects = await db.execute(select(Project).order_by(Project.id.desc()))
    return [ProjectOut.model_validate(p) for p in all_projects.scalars().all()]


@router.post("/{project_id}/share")
async def generate_share_link(project_id: int, db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    """Generate a share token for public access"""
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if not project.share_token:
        project.share_token = secrets.token_urlsafe(32)
        await db.flush()
        await db.refresh(project)

    return {"share_token": project.share_token}


@router.delete("/{project_id}/share")
async def revoke_share_link(project_id: int, db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    """Revoke share token"""
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    project.share_token = None
    await db.flush()
    return {"ok": True}


@router.get("/{project_id}", response_model=ProjectOut)
async def get_project(project_id: int, db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return ProjectOut.model_validate(project)


@router.put("/{project_id}", response_model=ProjectOut)
async def update_project(
    project_id: int, data: ProjectUpdate, db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)
):
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(project, k, v)
    await db.flush()
    await db.refresh(project)
    return ProjectOut.model_validate(project)


@router.delete("/{project_id}")
async def delete_project(project_id: int, db: AsyncSession = Depends(get_db), _: User = Depends(get_current_user)):
    result = await db.execute(select(Project).where(Project.id == project_id))
    project = result.scalar_one_or_none()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    await db.delete(project)
    return {"ok": True}
